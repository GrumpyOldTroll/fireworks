#!/usr/bin/env python3

from openpyxl import load_workbook, Workbook, styles
import sys
import argparse

LICENSE='''
Copyright 2925 Jake Holland  (jakeholland.net@gmail.com)

Redistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:

1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.

2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation and/or other materials provided with the distribution.

3. Neither the name of the copyright holder nor the names of its contributors may be used to endorse or promote products derived from this software without specific prior written permission.

THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS “AS IS” AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
'''

case_specs = '''
case counts, composed of racks of 5 guns of a caliber in a row:
A  BxC means:
  if you have A racks of a caliber: make B cases x C racks per case

1  1x1
2  1x2
3  1x3
4  1x4
5  1x5 preferred, 1x2 + 1x3 alternatively
6  2x3
7  1x3 + 1x4
8  1x4 + 1x4
9  3x3
10 2x3 + 1x4
11 1x3 + 2x4
12 3x4
'''

CALS={
  101: '4"',
  76: '3"',
  63: '2.5"',
}

CRATE_GROUPS={
  1: [1],
  2: [2],
  3: [3],
  4: [4],
  5: [2,3],
  6: [3,3],
  7: [3,4],
  8: [4,4],
  9: [3,3,3],
  10: [3,4,3],
  11: [3,4,4],
  12: [4,4,4],
}

PHASE_ORDERS=[
  [76,101,63],
  [101,76,63],
]

def cal_str(cal):
  if cal in CALS:
    return CALS[cal]
  return f'cal({cal})'

class CrateLayout(object):
  def __init__(self):
    self.cur_row = 1

  def write_layout(self, sheet, crates, cal, extras):
    cellid = 'A'+str(self.cur_row+2)
    sheet[cellid].value = cal_str(cal)

    col_off = 0
    for crate in crates:
      crate_rows = (len(crate)-1)//5 + 1
      for i in range(5):
        col_id = chr(ord('C')+col_off + i)
        for j in range(crate_rows):
          cellid = col_id + str(self.cur_row + j)
          pin = crate[i*crate_rows+j]
          sheet[cellid].value = f'{pin}({(pin-1)%50 + 1})' if pin > 50 else f'{pin}'

      col_off += 6

    if extras:
      col_id = chr(ord('C')+col_off) 
      for i, pin in enumerate(extras):
        cellid = col_id + str(self.cur_row + i)
        sheet[cellid].value = f'{pin}({(pin-1)%50+1})' if pin > 50 else f'{pin}'

    self.cur_row += 6

  def write_board(self, sheet, base):
    cellid = 'D'+str(self.cur_row)
    sheet[cellid].value = f'board(+{base})'
    side = lambda: styles.borders.Side(color='FF000000', style='thin')
    b = styles.borders.Border()
    b.top = side()
    b.bottom = side()
    b.left = side()
    sheet['A'+str(self.cur_row)].border = b
    for i in range(16):
      cellid = chr(ord('B')+i) + str(self.cur_row)
      b = styles.borders.Border()
      b.top = side()
      b.bottom = side()
      sheet[cellid].border = b
    sheet[chr(ord('A')+i)+str(self.cur_row)].border = b
    b = styles.borders.Border()
    b.top = side()
    b.bottom = side()
    b.right = side()

    self.cur_row += 2

class Board(object):
  def __init__(self):
    pass

  @staticmethod
  def make_border(sheet, row_str, start_idx, end_idx):
    side = lambda x: styles.borders.Side(color='FF000000', style=x)
    for i in range(start_idx, end_idx+1):
      cellid = chr(ord('B')+i) + row_str
      b = styles.borders.Border()
      if i == start_idx:
        if sheet[cellid].border.top is None or sheet[cellid].border.top.color is None:
          b.top = side('thin')
          b.bottom = side('thin')
        else:
          b.top = side('thick')
          b.bottom = side('thick')
          b.right = side('thin')
        b.left = side('thin')
      elif i == end_idx:
        b.top = side('thin')
        b.bottom = side('thin')
        b.right = side('thin')
      else:
        b.top = side('thin')
        b.bottom = side('thin')
      sheet[cellid].border = b

class KimBoard(Board):
  '''
layout looks like:
3" <count> <count> ...
+100	1	2	3	4	5	6	7	8	9	10	11	12	13	14	15	16	17	18	19	20	21	22	23	24	25	
	26	27	28	29	30	31	32	33	34	35	36	37	38	39	40	41	42	43	44	45	46	47	48	49	50	
4" <count> <count> ...
4" <count> <count> ...
	51	52	53	54	55	56	57	58	58	60	61	62	63	64	65	66	67	68	69	70	71	72	73	74	75	
	76	77	78	79	80	81	82	83	84	85	86	87	88	89	90	91	92	93	94	95	96	97	98	99	100	
3" <count> <count> ...

sample sheet with a row filled at:
https://docs.google.com/spreadsheets/d/1ydYXi8GyW79Vcrtl1Y7DAsNfZ3RV5xG4AiYxHMvmc5k/edit?gid=1202954332#gid=1202954332
'''

  def __init__(self):
    super().__init__()
    self.top_board_pin = None
    self.offset = None
    self.cal_pins = {}

  def add_row(self, pin, cal, qty):
    offset = ((pin - 1)//self.pincount())*self.pincount()
    if self.offset is None:
      self.offset = offset
      # NB: the board itself is 1-indexed, so the top pin is 50 or 100, not 49 or 99 --jake 2024-06-21
      self.top_board_pin = offset + self.pincount()

    elif self.offset != offset:
      raise ValueError(f'error: tried adding pin {pin} to KimBoard offset {self.offset}')

    idx = (pin - 1)%(self.pincount()//2)
    if cal not in self.cal_pins:
      self.cal_pins[cal] = [list() for _ in range(25)]
    self.cal_pins[cal][idx].extend([pin]*qty)

  def write_board_to_sheets(self, sheet, crate_sheet, crate_layout, phase, next_row):
    if self.offset % 100 == 0:
      if self.offset != 0:
        sheet['A'+str(next_row+1)] = f'+{self.offset}'
      pin_off = 0
    else:
      pin_off = 50

    for i in range(25):
      cellidx = chr(ord('B')+i) + str(next_row+1)
      sheet[cellidx] = i+1+pin_off
      cellidx = chr(ord('B')+i) + str(next_row+2)
      sheet[cellidx] = i+26+pin_off

    row_shifts = [0,3,4]  # same number as possible CAL values, how far to shift rows
    shift_idx = 0
    tail = 0
    wrote_layout_board = False
    for cal in PHASE_ORDERS[phase]:
      if cal not in self.cal_pins:
        continue
      row_str = str(next_row + row_shifts[shift_idx])
      tail = max(tail, row_shifts[shift_idx])
      shift_idx += 1
      sheet['A'+row_str] = CALS[cal]
      tot = 0
      pin_lists = self.cal_pins[cal]
      for i in range(25):
        val = len(pin_lists[i])
        if val:
          cellidx = chr(ord('B')+i) + row_str
          sheet[cellidx] = val
          tot += val
        sheet['AA'+row_str] = tot

      nfullracks = tot//5
      remainder = tot-nfullracks*5
      print(f'row {row_str}: {nfullracks} full racks, {remainder} extra')
      if remainder:
        sheet['AB'+row_str] = remainder

      if nfullracks not in CRATE_GROUPS:
        raise ValueError(f'row {row_str} has {tot} for {nfullracks} full racks, layout not in CRATE_GROUPS')

      crate_groups = CRATE_GROUPS[nfullracks]
      crates_list = []
      next_crate = -1
      crate_count = 0
      crate = None
      start_idx = None
      i = 0
      while i < len(pin_lists):
        pin_list = pin_lists[i]
        j = 0
        while j < len(pin_list):
          pin = pin_list[j]
          j += 1
          if crate is None:
            start_idx = i
            crate = []
            next_crate += 1
            if next_crate < len(crate_groups):
              crate_count = 5*crate_groups[next_crate]
            else:
              crate_count = 50
          crate.append(pin)
          if len(crate) >= crate_count:
            self.make_border(sheet, row_str, start_idx, i)
            crates_list.append(crate)
            start_idx = None
            crate = None

        i += 1

      crate_layout.write_layout(crate_sheet, crates_list, cal, crate)
      if not wrote_layout_board:
        wrote_layout_board = True
        crate_layout.write_board(crate_sheet, self.offset)

    return next_row + tail + 1

  @staticmethod
  def write_header(sheet, next_row):
    sheet['A'+str(next_row)] = "size"
    sheet['AA'+str(next_row)] = "count"
    sheet['AB'+str(next_row)] = "leftover"
    return next_row + 1

  @staticmethod
  def bt_name():
    return "Kim Slave"

  @staticmethod
  def pincount():
    return 50

class StripBoard(object):
  def __init__(self):
    self.pins = 50

class _LicenseAction(argparse.Action):
    def __init__(self,
                 option_strings,
                 license=None,
                 dest=argparse.SUPPRESS,
                 default=argparse.SUPPRESS,
                 help="show program's version number and exit"):
        super(_LicenseAction, self).__init__(
            option_strings=option_strings,
            dest=dest,
            default=default,
            nargs=0,
            help=help)
        self.license = license

    def __call__(self, parser, namespace, values, option_string=None):
        license = self.license
        if license is None:
            license = parser.license
        formatter = parser._get_formatter()
        formatter.add_text(license)
        parser.exit(message=formatter.format_help())

def main(in_args):
  argp = argparse.ArgumentParser(description="load an xlsx file with the fireworks show info, dump a plan for layout out crates made of racks.")
  argp.add_argument("--phased", "-p", help="flip the rows", action="store_true", default=False)
  argp.add_argument('--license', action=_LicenseAction, license=LICENSE, help='print license and exit')
  argp.add_argument('--version', action='version', help='print version', version='fireworks boards v0.0.1dev2024')
  argp.add_argument("fname", nargs=1, help="input file name (xlsx)")
  args = argp.parse_args(in_args[1:])

  fname = args.fname[0]

  print(f'loading {fname}')
  wb = load_workbook(fname)

  cue_sheet = wb.worksheets[0]
  # TBD: do we ever need to search thru multiple sheets or sheets that aren't the first? --jake 2024-06-21
  #for sheet_idx in range(len(wb.worksheets)):

  print(f'reading worksheet {cue_sheet.title}')

  pins, cals, qtys = None, None, None
  for col in cue_sheet.columns:
    if col[0].value == 'PIN':
      pins = [c.value for c in col[1:]]
    elif col[0].value == 'QTY':
      qtys = [c.value for c in col[1:]]
    elif col[0].value == 'CAL':
      cals = [c.value for c in col[1:]]

  errs = 0
  if pins is None:
    print(f'found no PIN column', file=sys.stderr)
    errs += 1
  if qtys is None:
    print(f'found no QTY column', file=sys.stderr)
    errs += 1
  if cals is None:
    print(f'found no CAL column', file=sys.stderr)
    errs += 1

  if errs != 0:
    print(f'problem scanning sheet "{cue_sheet.title}" from {fname}', file=sys.stderr)
    return -1

  board_types = [KimBoard]
  #board_types = [KimBoard, StripBoard]
  board_lists = []
  for board_type in board_types:
    cur_board = None
    board_list = []
    board_lists.append(board_list)
    for pin, qty, cal in zip(pins, qtys, cals):
      if cur_board is None or pin > cur_board.top_board_pin:
        cur_board = board_type()
        board_list.append(cur_board)
      cur_board.add_row(pin, cal, qty)

  outwb = Workbook()
  ws = None
  for board_type, board_list in zip(board_types, board_lists):
    if ws is None:
      ws = outwb.active
    else:
      ws = outwb.create_sheet(board_type.bt_name())
    ws.title = board_type.bt_name()
    ws.sheet_format.baseColWidth=6
    ws.sheet_format.defaultColWidth=6

    crate_layout = CrateLayout()
    crate_sheet = outwb.create_sheet(f'{board_type.bt_name()} layout')
    next_row = board_type.write_header(ws, 1)
    phase = 0 if not args.phased else 1
    for board in board_list:
      next_row = board.write_board_to_sheets(ws, crate_sheet, crate_layout, phase, next_row)
      phase = (phase + 1) % 2

  if args.phased:
    tail="_flipped"
  else:
    tail=""
  outfname = f'fireworks_boards{tail}.xlsx'
  print(f'writing {outfname}')
  outwb.save(outfname)

  return 0

if __name__=="__main__":
  ret = main(sys.argv)
  exit(ret)

